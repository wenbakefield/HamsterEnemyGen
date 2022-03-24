import sys
import os
import random
import time
import statistics
import openpyxl
import plotly.graph_objects as go
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib import style
import numpy as np
from numpy import diff
from collections import Counter

class Trait:
    def __init__(self, name, attack, defense, vulnerability):
        self.name = name
        self.attack = attack
        self.defense = defense
        self.vulnerability = vulnerability
    def __repr__(self):
        return "%s" % (self.name)
        # return "<Trait | Name:%s Attack:%s Defense:%s Vulnerability:%s>" % (self.name, self.attack, self.defense, self.vulnerability)
    def __str__(self):
        return "%s" % (self.name)
    def __eq__(self, other):
        if isinstance(other, Trait):
            return self.name == other.name and self.attack == other.attack and self.defense == other.defense and self.vulnerability == other.vulnerability

class Enemy:
    def __init__(self, species, trait, health):
        self.species = species
        self.trait = trait
        self.health = health
    def __repr__(self):
        return "<Enemy | Species:%s Trait:%s Health:%s = %s>" % (self.species, self.trait, self.health, sum(self.health) + self.trait.vulnerability)
    def __str__(self):
        return "%s %s with %s health" % (self.trait, self.species, sum(self.health) + self.trait.vulnerability)
    def __eq__(self, other):
        if isinstance(other, Enemy):
            return self.species == other.species and self.trait == other.trait and self.health == other.health

def get_random(pool):
  r, s = random.random(), 0
  for item in pool:
    s += item[1]
    if s >= r:
      return item[0]

def generate_enemies(species_list, traits_pool, health_pool, num_enemies):
    enemies = []
    count = 0
    while count < num_enemies:
        current_enemy_species = random.choice(species_list)
        current_enemy_trait = get_random(traits_pool)
        current_enemy_health = []
        current_enemy_health.append(get_random(health_pool))
        current_enemy_health.append(get_random(health_pool))
        current_enemy = Enemy(current_enemy_species, current_enemy_trait, current_enemy_health)
        enemies.append(current_enemy)
        count += 1
    return enemies

def get_enemy_fitness(enemy):
    global target_health_global
    fitness = abs(target_health_global - (sum(enemy.health) + enemy.trait.vulnerability))
    return fitness

def get_total_fitness(enemies):
    fitness = sum(map(get_enemy_fitness, enemies))
    return fitness

def sort_by_fitness(enemies):
    return sorted(enemies, key=get_enemy_fitness)

def is_fittest(enemy, cutoff):
    return get_enemy_fitness(enemy) <= cutoff

def get_fittest_enemies(enemies, cutoff):
    fittest_enemies = list(filter(lambda enemy: is_fittest(enemy, cutoff), enemies))
    return fittest_enemies

def get_percentage_fittest(enemies, cutoff):
    num_enemies = len(enemies)
    num_fittest_enemies = len(get_fittest_enemies(enemies, cutoff))
    return num_fittest_enemies / num_enemies

def get_popular_trait(enemies):
    traits = []
    for enemy in enemies:
        traits.append(enemy.trait.name)
    return statistics.mode(traits)

def get_popular_health_component(enemies):
    components = []
    for enemy in enemies:
        components.append(enemy.health[0])
        components.append(enemy.health[1])
    return statistics.mode(components)

def mutate_traits_pool(enemies, traits_pool):
    trait_names = []
    for enemy in enemies:
        trait_names.append(enemy.trait.name)
    trait_frequencies = Counter(trait_names)
    num_traits = len(trait_names)
    mutated_traits_pool = []
    for trait in traits_pool:
        trait_frequency = statistics.mean([(trait_frequencies[trait[0].name] / num_traits), trait[1]])
        mutated_traits_pool.append([trait[0], trait_frequency])
    return mutated_traits_pool

def mutate_health_pool(enemies, health_pool):
    health_components = []
    for enemy in enemies:
        health_components.append(enemy.health[0])
        health_components.append(enemy.health[1])
    health_component_frequencies = Counter(health_components)
    num_components = len(health_components)
    mutated_health_pool = []
    for num in health_pool:
        health_component_frequency = statistics.mean([(health_component_frequencies[num[0]] / num_components), num[1]])
        mutated_health_pool.append([num[0], health_component_frequency])
    return mutated_health_pool

# initialize pools
def species_list_init():
    return ["Bullfrog", 
            "Rat", 
            "Bat", 
            "Spider", 
            "Meerkat", 
            "Lizard", 
            "Snake", 
            "Rabbit", 
            "Owl", 
            "Pomeranian"]

def traits_pool_init():
    return [[Trait("Reserved", -1, 2, 0), 0.1],
            [Trait("Brave", 1, 0, -1), 0.1],
            [Trait("Reckless", 2, 0, -1), 0.1],
            [Trait("Cocky", 0, -2, 0), 0.1],
            [Trait("Buff", 1, 1, 2), 0.1],
            [Trait("Cheerful", 0, 0, 0), 0.1],
            [Trait("Lonely", -1, -1, 0), 0.1],
            [Trait("Desperate", 2, 0, -2), 0.1],
            [Trait("Weird", 0, 0, 0), 0.1],
            [Trait("Aloof", -1, 0, 1), 0.1]]

def health_pool_init():
    return [[1, 0.1],
            [2, 0.1],
            [3, 0.1],
            [4, 0.1],
            [5, 0.1],
            [6, 0.1],
            [7, 0.1],
            [8, 0.1],
            [9, 0.1],
            [10, 0.1]]

# initialize pools
species_list = species_list_init()
traits_pool = traits_pool_init()
health_pool = health_pool_init()

# initalize trackers
overall_fitness = 0
num_current_generation = 0
generation_reset_counter = 0
fitness_history = []
num_generation_history = []

# trait frequency tracking
reserved_freq = []
brave_freq = []
reckless_freq = []
cocky_freq = []
buff_freq = []
cheerful_freq = []
lonely_freq = []
desperate_freq = []
weird_freq = []
aloof_freq = []

# health frequency tracking
health_1_freq = []
health_2_freq = []
health_3_freq = []
health_4_freq = []
health_5_freq = []
health_6_freq = []
health_7_freq = []
health_8_freq = []
health_9_freq = []
health_10_freq = []

# initialize excel output
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(1, 1, "Generation")
ws.cell(1, 2, "Overall Fitness")
col = 3
for trait in traits_pool:
    ws.cell(1, col, trait[0].name)
    col += 1
for num in health_pool:
        ws.cell(1, col, num[0])
        col += 1

# initialize graphs

# main
print("Welcome to the Creature Evolver!")

target_health_global = int(input("Target Health: "))
generation_size = int(input("Generation Size: "))
num_kept_individuals = int(input("Number of Kept Individuals: "))
generation_limit = int(input("Maximum Generations: "))
desired_overall_fitness = int(input("Desired Overall Fitness (0 - 1): "))

while overall_fitness < desired_overall_fitness:

    num_generation_history.append(num_current_generation)
    current_generation = generate_enemies(species_list, traits_pool, health_pool, generation_size)
    overall_fitness = get_percentage_fittest(current_generation, 0)
    fitness_history.append(overall_fitness)
    
    fittest_individuals = sort_by_fitness(current_generation)
    fittest_individuals = fittest_individuals[:num_kept_individuals]
    best_trait = get_popular_trait(fittest_individuals)
    best_health_component = get_popular_health_component(fittest_individuals)

    print("Generation: ", end="")
    print(num_current_generation, end=" | ")

    print("Overall Fitness: ", end="")
    print("{:.3f}".format(overall_fitness), end=" | ")

    print("Best Health Componenent: ", end="")
    print(best_health_component, end=" | ")

    print("Best Trait: ", end="")
    print(best_trait)

    print("Trait Probabilities: ", end="")
    for trait in traits_pool:
        trait_name = trait[0].name
        trait_freq = trait[1]

        if trait_name == "Reserved":
            reserved_freq.append(trait_freq)
        if trait_name == "Brave":
            brave_freq.append(trait_freq)
        if trait_name == "Reckless":
            reckless_freq.append(trait_freq)
        if trait_name == "Cocky":
            cocky_freq.append(trait_freq)
        if trait_name == "Buff":
            buff_freq.append(trait_freq)
        if trait_name == "Cheerful":
            cheerful_freq.append(trait_freq)
        if trait_name == "Lonely":
            lonely_freq.append(trait_freq)
        if trait_name == "Desperate":
            desperate_freq.append(trait_freq)
        if trait_name == "Weird":
            weird_freq.append(trait_freq)
        if trait_name == "Aloof":
            aloof_freq.append(trait_freq)

        print(trait_name, end=" = ")
        print("{:.3f}".format(trait_freq, 3), end=" | ")
    print("")

    print("Health Probabilities: ", end="")
    for num in health_pool:
        health_name = num[0]
        health_freq = num[1]
        if health_name == 1:
            health_1_freq.append(health_freq)
        if health_name == 2:
            health_2_freq.append(health_freq)
        if health_name == 3:
            health_3_freq.append(health_freq)
        if health_name == 4:
            health_4_freq.append(health_freq)
        if health_name == 5:
            health_5_freq.append(health_freq)
        if health_name == 6:
            health_6_freq.append(health_freq)
        if health_name == 7:
            health_7_freq.append(health_freq)
        if health_name == 8:
            health_8_freq.append(health_freq)
        if health_name == 9:
            health_9_freq.append(health_freq)
        if health_name == 10:
            health_10_freq.append(health_freq)

        print(num[0], end=" = ")
        print("{:.3f}".format(num[1], 3), end=" | ")
    print("")


    # export to graphs

    # export to excel
    ws.cell(num_current_generation + 2, 1, num_current_generation)
    ws.cell(num_current_generation + 2, 2, overall_fitness)
    col = 3
    for trait in traits_pool:
        ws.cell(num_current_generation + 2, col, trait[1])
        col += 1
    for num in health_pool:
        ws.cell(num_current_generation + 2, col, num[1])
        col += 1

    traits_pool = mutate_traits_pool(fittest_individuals, traits_pool)
    health_pool = mutate_health_pool(fittest_individuals, health_pool)
    num_current_generation += 1
    generation_reset_counter += 1

    if num_current_generation >= 10:
        fitness_rate_of_change = statistics.mean(diff(fitness_history[num_current_generation - 10 : num_current_generation]))

        print("Fitness Rate of Change: ", end="")
        print("{:.3f}".format(fitness_rate_of_change))
        print("\n")

        print("Generation Reset Counter: ", end="")
        print(generation_reset_counter)
        print("\n")

        if abs(fitness_rate_of_change) < 0.001 and generation_reset_counter > 100:
            generation_reset_counter = 0
            traits_pool = traits_pool_init()
            health_pool = health_pool_init()

    if num_current_generation >= generation_limit:
        break

print("\n")
print("The Chosen Ones")
print("Generation: ", end="")
print(num_current_generation)
print("\n")

for enemy in current_generation:
    print(enemy)
print("\n")

wb.save("output.xlsx")
print("Data saved to output.xlsx")




"""
test_enemies = generate_enemies(species_list, traits_pool, health_pool, 100)
test_sorted = sort_by_fitness(test_enemies)
test_fittest = get_fit_enemies(test_enemies)

print(get_fit_percentage(test_enemies))
print(get_popular_trait(test_enemies))
print(get_popular_health_component(test_enemies))

for enemy in test_fittest:
    print(enemy)

for enemy in test_sorted:
    print(enemy)

"""